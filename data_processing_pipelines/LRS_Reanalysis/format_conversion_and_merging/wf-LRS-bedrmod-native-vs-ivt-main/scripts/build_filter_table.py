#!/usr/bin/env python3
"""
Assemble supplementary Excel table from per-biotype filter-count TSVs
produced by filter_sites.py --output-counts.

One sheet per biotype; columns derived from the TSV metadata header.
A "Total" row is appended at the bottom of each sheet.
"""

import argparse
import sys

import pandas as pd


def parse_args():
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument(
        "--dataset", nargs=2, metavar=("NAME", "COUNTS_TSV"),
        action="append", dest="datasets", default=[],
        help="Biotype name and counts TSV path (repeatable)",
    )
    p.add_argument("-o", "--output", required=True, help="Output .xlsx path")
    p.add_argument(
        "--dedup-counts", nargs=2, metavar=("NAME", "DEDUP_TSV"),
        action="append", dest="dedup_datasets", default=[],
        help="Biotype name and dedup counts TSV from proximity_filter.py (repeatable)",
    )
    p.add_argument(
        "--ivt-counts", nargs=2, metavar=("NAME", "IVT_TSV"),
        action="append", dest="ivt_datasets", default=[],
        help="Biotype name and total-IVT-sites counts TSV (repeatable)",
    )
    return p.parse_args()


def _display_columns(meta, short_cols):
    """Map short TSV column names to human-readable display labels."""
    mp = meta["max_padj"]
    lf = meta["min_log2fc"]
    sc = meta["min_ivt_absent_score"]

    fr = meta["min_native_freq"]
    mapping = {
        "total":        "Total Native sites",
        "ivt_testable": "IVT-testable",
        "ivt_absent":   "IVT-absent",
        f"padj_le_{mp}":
            f"padj ≤ {mp}",
        f"padj_le_{mp}_and_log2fc_gt_{lf}":
            f"padj ≤ {mp} & log2FC > {lf}",
        f"padj_le_{mp}_and_log2fc_gt_{lf}_freq_gt_{fr}":
            f"padj ≤ {mp} & log2FC > {lf}  &  freq > {fr}%",
        f"padj_le_{mp}_log2fc_gt_{lf}_or_score_gt_{sc}":
            f"(padj ≤ {mp} & log2FC > {lf}) OR score > {sc}",
        f"absent_and_score_gt_{sc}":
            f"IVT-absent & score > {sc}",
        f"padj_le_{mp}_log2fc_gt_{lf}_or_score_gt_{sc}_freq_gt_{fr}":
            f"(padj ≤ {mp} & log2FC > {lf}) OR score > {sc}  &  freq > {fr}%",
        f"absent_and_score_gt_{sc}_freq_gt_{fr}":
            f"IVT-absent & score > {sc}  &  freq > {fr}%",
        "after_filter": "After filter",
    }
    return [mapping.get(c, c) for c in short_cols]


def load_counts(path):
    """Return (meta_dict, DataFrame with modification as index)."""
    meta = {}
    with open(path) as fh:
        header_line = fh.readline()
        if header_line.startswith("#"):
            for token in header_line.lstrip("#").split():
                k, v = token.split("=")
                meta[k] = v
        else:
            raise ValueError(f"Expected metadata header line in {path}")

    df = pd.read_csv(path, sep="\t", comment="#")
    df = df.set_index("modification")
    return meta, df


def load_dedup_counts(path):
    """Return Series: modification → after_dedup count."""
    df = pd.read_csv(path, sep="\t", comment="#")
    return df.set_index("modification")["after_dedup"]


def load_ivt_counts(path):
    """Return Series: modification → total IVT sites."""
    df = pd.read_csv(path, sep="\t", comment="#")
    return df.set_index("modification")["ivt_total"]


def _style_sheet(worksheet, n_mods):
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    total_fill  = PatternFill("solid", fgColor="E2EFDA")
    bold        = Font(bold=True)

    for cell in worksheet[1]:
        cell.font      = bold
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    total_row = n_mods + 2
    for cell in worksheet[total_row]:
        cell.font = bold
        cell.fill = total_fill

    for row in worksheet.iter_rows():
        row[0].font = bold

    for col in worksheet.columns:
        width = max(len(str(cell.value or "")) for cell in col) + 2
        worksheet.column_dimensions[col[0].column_letter].width = min(width, 40)

    worksheet.row_dimensions[1].height = 60


def main():
    args = parse_args()
    if not args.datasets:
        sys.exit("error: provide at least one --dataset NAME COUNTS_TSV")

    dedup_by_name = {name: load_dedup_counts(path) for name, path in args.dedup_datasets}
    ivt_by_name   = {name: load_ivt_counts(path) for name, path in args.ivt_datasets}

    with pd.ExcelWriter(args.output, engine="openpyxl") as writer:
        for name, tsv_path in args.datasets:
            print(f"[info] Processing {name} ({tsv_path}) ...", file=sys.stderr)

            meta, df = load_counts(tsv_path)

            df.columns = _display_columns(meta, list(df.columns))
            df.index.name = "Modification"

            if name in ivt_by_name:
                df.insert(
                    0, "Total IVT sites",
                    ivt_by_name[name].reindex(df.index).fillna(0).astype(int),
                )

            if name in dedup_by_name:
                df["After dedup"] = dedup_by_name[name].reindex(df.index)

            total = df.sum(numeric_only=True)
            total.name = "Total"
            df = pd.concat([df, total.to_frame().T])

            n_mods = len(df) - 1
            df.to_excel(writer, sheet_name=name)
            _style_sheet(writer.sheets[name], n_mods)
            print(f"  {n_mods} modifications written", file=sys.stderr)

    print(f"[info] Written → {args.output}", file=sys.stderr)


if __name__ == "__main__":
    main()
